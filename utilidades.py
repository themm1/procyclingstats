import openpyxl
from openpyxl.styles import Font, Alignment
import os
import pandas as pd
#import polars as pl
from pandas import ExcelWriter
import gc
import requests
from io import BytesIO
from openpyxl import load_workbook,Workbook,worksheet,utils
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from procyclingstats import Race, Rider, Stage
from procyclingstats import RaceClimbs
from openpyxl.drawing.image import Image

from PIL import Image as PILImage

# carga la lista de carreras desde un fichero Excel alojado en Dropbox
def cargar_lista_carreras():
    
    # Increase the decompression bomb limit for large Excel files
    PILImage.MAX_IMAGE_PIXELS = None
    
    carrera=pd.DataFrame(columns=["carrera","url"])
    url="https://www.dropbox.com/scl/fi/k9vozw3agqbbdx6si4rvx/Burgos-Burpellet-BH-26.xlsx?rlkey=j8ch74vpx7g82bx4ee3bfho3b&dl=1"
    #url=("C://Users//echav//Dropbox//DIRECCIÓN BBH26//CARRERAS 2026//Burgos-Burpellet-BH 26.xlsx")
    try:
       
        # Check if URL is a web URL or local file path
        if url.startswith(('http://', 'https://')):
            # For web URLs, use requests
            response = requests.get(url)
            response.raise_for_status()  
            # para asegurar descarga correcta
            archivo_memoria = BytesIO(response.content)
            wb=openpyxl.load_workbook(archivo_memoria)
        else:
            # For local file paths, load directly
            wb=openpyxl.load_workbook(url)
        print("Fichero cargado correctamente.")
        ws = wb.active
        
        if wb and ws:
                for row in ws.iter_rows(min_col=2, min_row=1, max_row=5, values_only=False):
                    for celda in row:
                        if celda.hyperlink:
                            carrera.loc[len(carrera)] = [celda.value,str(celda.hyperlink.target).split(".com/")[-1]]
                            #print(f"{celda.value}: {celda.hyperlink.target}")
        
    except FileNotFoundError:
        print(f"El fichero '{url}' no existe.")
       
    except Exception as e:
        print(f"Error al cargar el fichero: {e}")    
    return carrera

def find_logo_path(filename: str = "logo.png"):
    """Intenta resolver la ruta del logo.
    Prioriza env var ``PCS_LOGO_PATH`` y la ruta estándar ``web_carreras/assets/logo.png``.
    Devuelve la ruta absoluta si existe, o None si no se encuentra.
    """
    # 1) Variable de entorno
    env_path = os.getenv("PCS_LOGO_PATH")
    if env_path and os.path.exists(env_path):
        return os.path.abspath(env_path)

    # 2) Ubicaciones relativas comunes
    try:
        file_dir = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        file_dir = os.getcwd()
    cwd = os.getcwd()

    standard_rel = os.path.join("web_carreras", "assets", filename)
    candidates = [
        # Ubicación estándar dentro del proyecto
        os.path.join(file_dir, standard_rel),
        os.path.join(cwd, standard_rel),
        # Compatibilidad retro: raíz del proyecto
        os.path.join(file_dir, filename),
        os.path.join(cwd, filename),
        # Compatibilidad retro: web_carreras/ directo
        os.path.join(file_dir, "web_carreras", filename),
        os.path.join(cwd, "web_carreras", filename),
    ]
    for path in candidates:
        if os.path.exists(path):
            return os.path.abspath(path)
    return None

def buscar_resultados_carrera(enlace_o_valor):
    """
    Intenta obtener resultados de carrera usando procyclingstats.Race.
    entrada: puede ser una URL completa, un slug o None.
    devuelve: DataFrame con resultados o None si falla.
    """
    if not enlace_o_valor:
        print("No se proporcionó enlace o valor.")
        return None
      
       
    try:
        print("Intentando obtener resultados para:", enlace_o_valor)
        race= Race(f"{enlace_o_valor}/overview")     
        df_res = pd.DataFrame()  
        df_fin = pd.DataFrame()
        #df_excell = pd.DataFrame()        
        print("Obteniendo:"+race.name())
        if race.is_one_day_race(): 
            df_res =one_day_race_results(enlace_o_valor)
            return df_res
        #sacamos los resultados de los 5 primeras de cada etapa
        stages = race.stages() 
        
        book = load_workbook(f"{race.name()}_results.xlsx")
        with ExcelWriter(f"{race.name()}_results.xlsx",mode='a',if_sheet_exists ='overlay') as writer:
            for stage_info in stages:
                stage = Stage(stage_info['stage_url'])
                print("Obteniendo resultados de la etapa:", stage.relative_url())
                # Obtener resultados de la etapa
                df_res =pd.DataFrame(stage.parse().get('results'))
                df_fin=pd.concat([df_fin, df_res.head(10)])  
                #df_res.head(10).to_excel(f"{race.name()}_results.xlsx")
                df_res.head().to_excel(writer,index=False)  
                df_res[['rider_name', 'team_name','rank','uci_points']].head().to_excel(writer)
                
        return df_fin[['rider_name', 'team_name','rank','uci_points']]  
           
    except Exception as e:
        print("Error al obtener resultados de la carrera:", e)
        return None 
def one_day_race_results(race_slug):
    race = Race(f"{race_slug}/overview") 
    file=f"analisis_carrera_{race.name()}.xlsx"

    libro=Workbook()
    hojas=libro.active  

    df_res = pd.DataFrame()  
    df_fin = pd.DataFrame()
    df_rider_specialidad = pd.DataFrame()
    list_editions = race.prev_editions_select() 
    c=0
    logo_path = find_logo_path('logo.png')
    if logo_path:
        try:
            img = Image(logo_path)
            hojas.add_image(img, 'A1')
            hojas.merge_cells('A1:C6')
            img.height = 170
            img.width = 410
        except Exception as e:
            print(f"No se pudo cargar el logo desde '{logo_path}': {e}")
    else:
        print("Logo no encontrado; continuando sin imagen.")
    hojas.title=race.name() # type: ignore
    # Fusionar A1:E1 y aplicar fuente negra tamaño 16
    hojas['D4'] = race.name()+" "+race.uci_tour()# type: ignore
    hojas.merge_cells('D4:G4')
    hojas['D4'].font = Font(color="F40BE4", size=30, bold=True) 
    hojas['D4'].alignment = Alignment(horizontal="center", vertical="center")
    hojas[f'A{hojas.max_row+2}']=f"EDICION:{race.year() }, Fecha:{race.startdate()}"
    hojas[f'A{hojas.max_row}'].font = Font(color="000000", size=25, bold=True)
    #obtenemos informacion de los puertos ediciaon actual
    race_climbs = RaceClimbs(f"{race_slug}/route/climbs")
    if race_climbs is not None:
        df_climbs_current = pd.DataFrame(race_climbs.climbs())
    #hojas[f'A{hojas.max_row}']=f"Puertos edición"
    end_row_res = write_df_to_sheet(
        hojas, df_climbs_current, start_row=hojas.max_row+1, start_col=1,
        table_name=f"Climbs_{race.name().replace('-', '_')}",
        style_name="TableStyleMedium5",
        add_table=True
    )
    if(len(list_editions)>5):
        cont=6
    else:
        cont=len(list_editions)
    for edition in list_editions[1:cont]:
   
        print("Procesando edición", edition['text'])
        hojas[f'A{hojas.max_row+2}']=f"Edición {edition['text']}"
        hojas[f'A{hojas.max_row}'].font = Font(color="000000", size=16, bold=True)
    
  
        past_edit=Stage(f"{race_slug.strip()[0:len(race_slug)-4]}{edition['text']}/result")
        race_climbs = RaceClimbs(f"{race_slug.strip()[0:len(race_slug)-4]}{edition['text']}/route/climbs")
        df_climbs = pd.DataFrame(race_climbs.climbs()) if race_climbs is not None else pd.DataFrame()
        if not df_climbs.empty:
            df_climbs.rename(columns={'climb_name': 'Nombre', 'length': 'kms', 'steepness': 'Media','top':'altitud','km_before_finnish':'Distancia a Meta'}, inplace=True)
            hojas[f'A{hojas.max_row+1}']=f"Puertos edición"
            end_row_res=write_df_to_sheet(
            hojas, df_climbs[['Nombre', 'kms','Media','altitud','Distancia a Meta']].sort_values(by='Distancia a Meta', ascending=False), start_row=hojas.max_row+1, start_col=1,
            table_name=f"climbs_{edition['text'].replace('-', '_')}",
            style_name="TableStyleMedium2",
            add_table=True
            )
        print(f'como:{hojas.max_row},participacion:{past_edit.race_startlist_quality_score()} ,desnivel: {past_edit.vertical_meters()}m ,distancia {past_edit.distance()}km, avg:{past_edit.avg_speed_winner()}km/h')
        c=hojas.max_row+2
        hojas[f'A{c-1}']="Resumen:"
        hojas[f'A{c+1}']=f'{past_edit.won_how()}'
        hojas[f'A{c+1}'].font = Font(color="FF0000", bold=True)
        hojas[f'B{c+1}']=f'Participacion:{past_edit.race_startlist_quality_score()}'
        hojas[f'C{c+1}']=f'Desnivel: {past_edit.vertical_meters()}m'# type: ignore
        hojas[f'D{c+1}']=f'{past_edit.distance()}km'# type: ignore
        hojas[f'E{c+1}']=f'{past_edit.avg_speed_winner()}km/h'# type: ignore
    
        df_res = pd.DataFrame(past_edit.parse()['results'])# type: ignore
    
        if(df_res.empty):
            print("  No hay resultados para esta edición.")
            hojas.append(["  No hay resultados para esta edición."])# type: ignore
            continue
    
        # Obtener especialidad de los 10 primeros ciclistas
        df_riders_specialties = []
        for idx, row in df_res.head(10).iterrows():
            try:
                rider = Rider(str(row['rider_url']))
                rider_data = rider.parse()
                points_per_speciality = rider_data.get('points_per_speciality', {})
            
                if isinstance(points_per_speciality, dict) and points_per_speciality:
                    sorted_by_values = dict(sorted(points_per_speciality.items(), key=lambda item: item[1], reverse=True))
                    # Tomar las 2 primeras especialidades
                    top_2 = list(sorted_by_values.items())[:2]
                    rider_info = {                    
                    'specialty_1': top_2[0][0] if len(top_2) > 0 else '',
                    'points_1': top_2[0][1] if len(top_2) > 0 else 0,
                    'specialty_2': top_2[1][0] if len(top_2) > 1 else '',
                    'points_2': top_2[1][1] if len(top_2) > 1 else 0
                    }
                    df_riders_specialties.append(rider_info)
            except Exception as e:
                print(f"Error al procesar ciclista {row.get('rider_name', 'unknown')}: {e}")
    
        df_rider_specialidad = pd.DataFrame(df_riders_specialties)
    
        # Preparar tabla de resultados con especialidades
        df_res_subset = df_res[['rank', 'rider_name','time','team_name']].head(10).reset_index(drop=True)
        specialty_series = pd.Series([""] * len(df_res_subset))
        if not df_rider_specialidad.empty:
            specialty_series = df_rider_specialidad.apply(
            lambda r: f"{r['specialty_1']}:{r['points_1']}, {r['specialty_2']}:{r['points_2']}",
            axis=1
            ).reindex(df_res_subset.index).fillna("")
        df_res_subset['specialties'] = specialty_series
    
        #sacamos ciclistas Burgos y puntos UCI 
        df_burgos = df_res[df_res['team_name'].str.contains('Burgos')]
        df_uci=df_res.groupby('team_name')['uci_points'].sum().reset_index().sort_values(by='uci_points', ascending=False)
    
        # Escribir df_res con especialidades como última columna
        hojas[f'A{hojas.max_row+2}']="TOP 10"
        end_row_res = write_df_to_sheet(
            hojas, df_res_subset, start_row=hojas.max_row+2, start_col=1,
            table_name=f"Resultados_{edition['text'].replace('-', '_')}",
            style_name="TableStyleMedium5",
            add_table=True
            )
    
        # Escribir df_burgos debajo de df_res (A-E) sin tabla
        end_row_burgos = end_row_res
        if not df_burgos.empty:
            df_burgos_to_write = df_burgos[['rank', 'rider_name','time','uci_points','breakaway_kms']]
            hojas[f'A{hojas.max_row+2}']="RESULTADO BURGOS"
            end_row_burgos = write_df_to_sheet(
                hojas, df_burgos_to_write, start_row=end_row_res+3, start_col=1,
                table_name=f"Burgos_{edition['text'].replace('-', '_')}",
                style_name="TableStyleMedium5",
                show_first_col=True, show_col_stripes=True,
                add_table=True
                )
    
        # Escribir df_uci (G-H) sin tabla
        df_uci_to_write = df_uci
        hojas[f'G{c+2}']="PUNTOS UCI POR EQUIPO"
        end_row_uci = write_df_to_sheet(
            hojas, df_uci_to_write, start_row=c+3, start_col=7,
            table_name=f"UCI_{edition['text'].replace('-', '_')}",
            style_name="TableStyleMedium5",
            show_first_col=True,
            add_table=True
            )
        gc.collect() 
    
    libro.save (file)
    df_final = pd.DataFrame(libro.active.values)
    print(f"Análisis guardado en {df_final.dropna().head()}")
    return(df_final.dropna())

def write_df_to_sheet(ws, df, start_row, start_col, table_name, style_name="TableStyleMedium9",
                      show_first_col=False, show_last_col=False, show_row_stripes=True, show_col_stripes=False,
                      add_table=False, print_to_console=True):
    if df is None or df.empty:
        return start_row
    
    # Opcional: imprimir DataFrame en pantalla
    if print_to_console:
        print(f"\n=== {table_name} ({len(df)} filas) ===")
        print(df.to_string(index=False))
        print("")
    
    # Encabezados
    for i, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=i, value=str(col))
    # Datos
    for r_i, row in enumerate(df.itertuples(index=False, name=None), start=start_row + 1):
        for c_i, value in enumerate(row, start=start_col):
            ws.cell(row=r_i, column=c_i, value=value)
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    if add_table:
        ref = f"{utils.get_column_letter(start_col)}{start_row}:{utils.get_column_letter(end_col)}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=style_name,
            showFirstColumn=show_first_col,
            showLastColumn=show_last_col,
            showRowStripes=show_row_stripes,
            showColumnStripes=show_col_stripes
        )
        ws.add_table(table)
    return end_row



def write_df_to_sheet_especialidad(ws, df, start_row, start_col):
    """Escribe cada fila del DataFrame de especialidades en una sola celda"""
    if df is None or df.empty:
        return start_row
    
    current_row = start_row
    max_width = 0
    
    # Procesar cada fila del DataFrame
    for idx, row in df.iterrows():
        # Concatenar las especialidades en una sola cadena
        cell_value = f"{row['specialty_1']}:{row['points_1']}, {row['specialty_2']}:{row['points_2']}"
        ws.cell(row=current_row, column=start_col, value=cell_value)
        max_width = max(max_width, len(cell_value))
        current_row += 1
    
    # Ajustar ancho de columna
    col_letter = utils.get_column_letter(start_col)
    ws.column_dimensions[col_letter].width = max(max_width + 2, ws.column_dimensions[col_letter].width or 0)
    
    return current_row


if __name__ == "__main__": 
    cargar_lista_carreras()
    #buscar_resultados_carrera("https://www.procyclingstats.com/race/gran-premio-miguel-indurain/2026")